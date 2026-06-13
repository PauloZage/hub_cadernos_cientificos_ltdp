# ============================================================================
# HUB DE CADERNOS CIENTÍFICOS — LTDP (Lubango, Huíla, Angola)
# Backend API — FastAPI + SQLAlchemy (SQLite por omissão, PostgreSQL via env)
# Fuso horário oficial: WAT — West Africa Time (UTC+1) / Africa/Luanda
# ----------------------------------------------------------------------------
# Execução:
#   pip install -r requirements.txt
#   uvicorn main:app --host 0.0.0.0 --port 8000
# Documentação interactiva: http://localhost:8000/docs
# ============================================================================

import hashlib
import hmac
import io
import json
import os
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional

import jwt
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel, Field
from sqlalchemy import (Column, ForeignKey, Integer, String, Text,
                        create_engine, func)
from sqlalchemy.orm import Session, declarative_base, sessionmaker

# ----------------------------------------------------------------------------
# Configuração e contexto local
# ----------------------------------------------------------------------------
LOCALIZACAO = "LTDP - Lubango, Huíla, Angola"
FUSO_LABEL = "WAT (UTC+1)"
WAT = timezone(timedelta(hours=1), name="WAT")

# Render injecta DATABASE_URL automaticamente quando há um Postgres ligado ao serviço
DATABASE_URL = os.getenv("LTDP_DATABASE_URL") or os.getenv("DATABASE_URL") or "sqlite:///./ltdp_hub.db"
if DATABASE_URL.startswith("postgres://"):  # SQLAlchemy 2.x exige o prefixo completo
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
JWT_SECRET = os.getenv("LTDP_JWT_SECRET", "ALTERAR-EM-PRODUCAO-lubango-huila")
JWT_ALG = "HS256"
JWT_EXPIRA_HORAS = 12  # sessões longas: reduz re-autenticações em redes instáveis

engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {},
)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()


def agora_wat() -> datetime:
    return datetime.now(WAT)


# ----------------------------------------------------------------------------
# Modelos ORM
# ----------------------------------------------------------------------------
class Utilizador(Base):
    __tablename__ = "utilizadores"
    id_utilizador = Column(String, primary_key=True)
    nome_completo = Column(String, nullable=False)
    email = Column(String, unique=True, nullable=False)
    perfil = Column(String, nullable=False, default="INVESTIGADOR")
    senha_hash = Column(String, nullable=False)
    activo = Column(Integer, nullable=False, default=1)
    criado_em = Column(String, nullable=False)


class Caderno(Base):
    __tablename__ = "cadernos"
    id_caderno = Column(String, primary_key=True)
    codigo_ltdp = Column(String, unique=True, nullable=False)
    titulo_invencao = Column(String, nullable=False)
    investigador_principal = Column(String, nullable=False)
    # NOVO: outros participantes na invenção (texto livre: nomes separados por vírgula ou linha)
    participantes_adicionais = Column(Text, nullable=True)
    # NOVO: número telefónico de contacto do caderno/invenção
    telefone_contacto = Column(String, nullable=True)
    area_tecnologica = Column(String, nullable=False)
    data_inicio = Column(String, nullable=False)
    status_trl = Column(Integer, nullable=False, default=1)
    codigo_patente_previsto = Column(String, nullable=True)
    estado = Column(String, nullable=False, default="ACTIVO")
    localizacao = Column(String, nullable=False, default=LOCALIZACAO)
    fuso_horario = Column(String, nullable=False, default=FUSO_LABEL)
    criado_em = Column(String, nullable=False)
    actualizado_em = Column(String, nullable=False)


class ImagemInvento(Base):
    """NOVO: imagens dos inventos, carregadas a partir do PC do utilizador."""
    __tablename__ = "imagens_invento"
    id_imagem = Column(String, primary_key=True)
    id_caderno = Column(String, ForeignKey("cadernos.id_caderno"), nullable=False)
    nome_ficheiro = Column(String, nullable=False)
    tipo_mime = Column(String, nullable=False)
    legenda = Column(String, nullable=True)
    dados_base64 = Column(Text, nullable=False)   # imagem embebida (base64) — sem dependência de storage externo
    carregado_por = Column(String, nullable=False)
    criado_em = Column(String, nullable=False)


class CustoInvento(Base):
    """NOVO: custos da invenção, do início à finalização (rastreio financeiro)."""
    __tablename__ = "custos_invento"
    id_custo = Column(String, primary_key=True)
    id_caderno = Column(String, ForeignKey("cadernos.id_caderno"), nullable=False)
    descricao = Column(String, nullable=False)
    categoria = Column(String, nullable=False, default="MATERIAL")  # MATERIAL, EQUIPAMENTO, MAO_OBRA, SERVICOS, OUTROS
    valor = Column(String, nullable=False)        # guardado como texto para preservar precisão (Kwanza)
    moeda = Column(String, nullable=False, default="AOA")
    data_despesa = Column(String, nullable=False)  # YYYY-MM-DD
    fase = Column(String, nullable=False, default="DESENVOLVIMENTO")  # CONCEPCAO, DESENVOLVIMENTO, PROTOTIPO, TESTES, FINALIZACAO
    registado_por = Column(String, nullable=False)
    criado_em = Column(String, nullable=False)


class LogSincronizacao(Base):
    __tablename__ = "log_sincronizacao"
    id_sync = Column(String, primary_key=True)
    id_utilizador = Column(String, ForeignKey("utilizadores.id_utilizador"), nullable=True)
    data_sync = Column(String, nullable=False)
    entradas_novas = Column(Integer, nullable=False, default=0)
    duplicados = Column(Integer, nullable=False, default=0)
    alertas_hash = Column(Integer, nullable=False, default=0)
    ficheiro_origem = Column(String, nullable=True)
    detalhes = Column(Text, nullable=True)


class EntradaCientifica(Base):
    __tablename__ = "entradas_cientificas"
    id_entrada = Column(String, primary_key=True)
    id_caderno = Column(String, ForeignKey("cadernos.id_caderno"), nullable=False)
    data_registo = Column(String, nullable=False)
    metodologia = Column(Text, nullable=False)
    resultados_brutos = Column(Text, nullable=False)
    link_repositorio_codigo = Column(String, nullable=True)
    hash_seguranca = Column(String, unique=True, nullable=False)
    hash_anterior = Column(String, nullable=True)
    assinatura_digital_investigador = Column(String, nullable=False)
    assinatura_testemunha = Column(String, nullable=True)
    origem_registo = Column(String, nullable=False, default="API")
    localizacao = Column(String, nullable=False, default=LOCALIZACAO)
    fuso_horario = Column(String, nullable=False, default=FUSO_LABEL)
    entrada_corrigida = Column(String, nullable=True)
    criado_em = Column(String, nullable=False)
    # NOVO: campos de auditoria para edição/anulação pelo Diretor (administrador)
    anulada = Column(Integer, nullable=False, default=0)        # soft-delete: 1 = anulada
    editada_em = Column(String, nullable=True)
    editada_por = Column(String, nullable=True)
    motivo_alteracao = Column(Text, nullable=True)


Base.metadata.create_all(engine)


# ----------------------------------------------------------------------------
# Migração leve: adiciona colunas novas a bases de dados já existentes.
# (create_all cria tabelas em falta mas NÃO altera tabelas já criadas.)
# ----------------------------------------------------------------------------
def _migrar_colunas():
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if "sqlite" in DATABASE_URL:
        tipo = "TEXT"
    else:
        tipo = "TEXT"
    novas = {
        "cadernos": [("participantes_adicionais", tipo), ("telefone_contacto", tipo)],
        "entradas_cientificas": [
            ("anulada", "INTEGER DEFAULT 0"), ("editada_em", tipo),
            ("editada_por", tipo), ("motivo_alteracao", tipo),
        ],
    }
    with engine.begin() as conn:
        for tabela, colunas in novas.items():
            try:
                existentes = {c["name"] for c in insp.get_columns(tabela)}
            except Exception:
                continue
            for nome_col, tipo_col in colunas:
                if nome_col not in existentes:
                    try:
                        conn.execute(text(f'ALTER TABLE {tabela} ADD COLUMN {nome_col} {tipo_col}'))
                    except Exception:
                        pass


_migrar_colunas()

# ----------------------------------------------------------------------------
# Integridade intelectual: SHA-256 canónico e assinaturas
# ----------------------------------------------------------------------------
def calcular_hash_entrada(id_caderno: str, data_registo: str, metodologia: str,
                          resultados_brutos: str, link_repo: Optional[str]) -> str:
    """
    Hash canónico da entrada. A MESMA fórmula é usada pelo script de
    sincronização Excel — qualquer divergência indica adulteração dos dados.
    """
    payload = json.dumps(
        {
            "id_caderno": id_caderno,
            "data_registo": data_registo,
            "metodologia": metodologia.strip(),
            "resultados_brutos": resultados_brutos.strip(),
            "link_repositorio_codigo": (link_repo or "").strip(),
            "contexto": LOCALIZACAO,
        },
        ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def assinar(email: str, hash_seguranca: str) -> str:
    """Assinatura digital leve: HMAC-SHA256(segredo do servidor, email + hash)."""
    msg = f"{email}|{hash_seguranca}".encode("utf-8")
    return hmac.new(JWT_SECRET.encode("utf-8"), msg, hashlib.sha256).hexdigest()


def hash_senha(senha: str, salt: Optional[str] = None) -> str:
    salt = salt or uuid.uuid4().hex
    dk = hashlib.pbkdf2_hmac("sha256", senha.encode(), salt.encode(), 120_000)
    return f"{salt}${dk.hex()}"


def verificar_senha(senha: str, guardado: str) -> bool:
    salt, _ = guardado.split("$", 1)
    return hmac.compare_digest(hash_senha(senha, salt), guardado)


# ----------------------------------------------------------------------------
# Autenticação JWT (leve, sem dependências pesadas)
# ----------------------------------------------------------------------------
oauth2 = OAuth2PasswordBearer(tokenUrl="auth/login")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def utilizador_actual(token: str = Depends(oauth2), db: Session = Depends(get_db)) -> Utilizador:
    erro = HTTPException(status.HTTP_401_UNAUTHORIZED, "Credenciais inválidas ou sessão expirada.")
    try:
        dados = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise erro
    u = db.get(Utilizador, dados.get("sub", ""))
    if not u or not u.activo:
        raise erro
    return u


def exigir_diretor(u: Utilizador = Depends(utilizador_actual)) -> Utilizador:
    if u.perfil != "DIRETOR":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Acesso reservado ao Diretor do LTDP.")
    return u


# ----------------------------------------------------------------------------
# Schemas Pydantic
# ----------------------------------------------------------------------------
class CadernoIn(BaseModel):
    titulo_invencao: str
    investigador_principal: str
    participantes_adicionais: Optional[str] = Field(None, description="Outros nomes participantes na invenção.")
    telefone_contacto: Optional[str] = Field(None, description="Número telefónico de contacto.")
    area_tecnologica: str
    data_inicio: str = Field(..., description="YYYY-MM-DD")
    status_trl: int = Field(1, ge=1, le=9)
    codigo_patente_previsto: Optional[str] = None
    id_caderno: Optional[str] = Field(None, description="UUID gerado offline (sync Excel). Se omitido, o Hub gera.")


class EntradaEditar(BaseModel):
    """NOVO: edição de entrada já gravada (reservado ao Diretor)."""
    metodologia: Optional[str] = None
    resultados_brutos: Optional[str] = None
    link_repositorio_codigo: Optional[str] = None
    data_registo: Optional[str] = None
    motivo_alteracao: str = Field(..., min_length=3, description="Justificação obrigatória da correcção (auditoria).")


class CustoIn(BaseModel):
    """NOVO: registo de um custo da invenção."""
    descricao: str
    categoria: str = Field("MATERIAL", pattern="^(MATERIAL|EQUIPAMENTO|MAO_OBRA|SERVICOS|OUTROS)$")
    valor: float = Field(..., ge=0)
    moeda: str = Field("AOA")
    data_despesa: str = Field(..., description="YYYY-MM-DD")
    fase: str = Field("DESENVOLVIMENTO", pattern="^(CONCEPCAO|DESENVOLVIMENTO|PROTOTIPO|TESTES|FINALIZACAO)$")


class EntradaIn(BaseModel):
    data_registo: Optional[str] = Field(None, description="ISO 8601; se omitido, usa agora em WAT.")
    metodologia: str
    resultados_brutos: str
    link_repositorio_codigo: Optional[str] = None
    assinatura_testemunha_email: Optional[str] = None
    id_entrada: Optional[str] = Field(None, description="UUID gerado offline (sync Excel) → idempotência.")
    origem_registo: str = Field("API", pattern="^(API|EXCEL_SYNC)$")


class UtilizadorIn(BaseModel):
    nome_completo: str
    email: str
    senha: str
    perfil: str = Field("INVESTIGADOR", pattern="^(DIRETOR|INVESTIGADOR|TESTEMUNHA)$")


# ----------------------------------------------------------------------------
# Aplicação
# ----------------------------------------------------------------------------
app = FastAPI(
    title="Hub de Cadernos Científicos — LTDP",
    description=f"Ecossistema de Cadernos de Laboratório Electrónicos (ELN). {LOCALIZACAO} — {FUSO_LABEL}.",
    version="1.1.0",
)

# CORS: permite que o script de sincronização e futuras apps acedam de outras origens
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("LTDP_CORS_ORIGINS", "*").split(","),
    allow_methods=["*"], allow_headers=["*"],
)

PASTA_RAIZ = os.path.dirname(os.path.abspath(__file__))  # index.html vive ao lado de main.py


@app.on_event("startup")
def criar_diretor_inicial():
    """Cria o utilizador Diretor por omissão se a base estiver vazia."""
    db = SessionLocal()
    try:
        if db.query(Utilizador).count() == 0:
            db.add(Utilizador(
                id_utilizador=str(uuid.uuid4()),
                nome_completo="Diretor LTDP",
                email=os.getenv("LTDP_DIRETOR_EMAIL", "diretor@ltdp.ao"),
                perfil="DIRETOR",
                senha_hash=hash_senha(os.getenv("LTDP_DIRETOR_SENHA", "ltdp2026")),
                criado_em=agora_wat().isoformat(),
            ))
            db.commit()
    finally:
        db.close()


@app.get("/", tags=["Sistema"], include_in_schema=False)
def interface_web():
    """Serve a interface web do Hub (login, painel do Diretor, cadernos)."""
    return FileResponse(os.path.join(PASTA_RAIZ, "index.html"))


@app.get("/health", tags=["Sistema"])
def estado_sistema():
    return {"sistema": "Hub de Cadernos Científicos — LTDP", "localizacao": LOCALIZACAO,
            "fuso_horario": FUSO_LABEL, "hora_local": agora_wat().isoformat(), "estado": "online"}


# ---------------------------- Autenticação ---------------------------------
@app.post("/auth/login", tags=["Autenticação"])
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    u = db.query(Utilizador).filter(Utilizador.email == form.username).first()
    if not u or not verificar_senha(form.password, u.senha_hash):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Email ou senha incorrectos.")
    token = jwt.encode(
        {"sub": u.id_utilizador, "perfil": u.perfil,
         "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRA_HORAS)},
        JWT_SECRET, algorithm=JWT_ALG,
    )
    return {"access_token": token, "token_type": "bearer", "perfil": u.perfil,
            "expira_em_horas": JWT_EXPIRA_HORAS}


@app.post("/auth/utilizadores", tags=["Autenticação"], status_code=201)
def criar_utilizador(dados: UtilizadorIn, db: Session = Depends(get_db),
                     _: Utilizador = Depends(exigir_diretor)):
    if db.query(Utilizador).filter(Utilizador.email == dados.email).first():
        raise HTTPException(409, "Já existe um utilizador com este email.")
    u = Utilizador(id_utilizador=str(uuid.uuid4()), nome_completo=dados.nome_completo,
                   email=dados.email, perfil=dados.perfil,
                   senha_hash=hash_senha(dados.senha), criado_em=agora_wat().isoformat())
    db.add(u); db.commit()
    return {"id_utilizador": u.id_utilizador, "email": u.email, "perfil": u.perfil}


# ------------------------------ Cadernos ------------------------------------
@app.post("/cadernos", tags=["Cadernos"], status_code=201)
def abrir_caderno(dados: CadernoIn, db: Session = Depends(get_db),
                  u: Utilizador = Depends(utilizador_actual)):
    """Abre um novo caderno de invenção no Hub."""
    cid = dados.id_caderno or str(uuid.uuid4())
    if db.get(Caderno, cid):
        raise HTTPException(409, "Caderno já existe no Hub (sincronização idempotente).")
    ano = agora_wat().year
    seq = db.query(func.count(Caderno.id_caderno)).scalar() + 1
    agora = agora_wat().isoformat()
    c = Caderno(
        id_caderno=cid, codigo_ltdp=f"LTDP/CAD/{ano}/{seq:04d}",
        titulo_invencao=dados.titulo_invencao,
        investigador_principal=dados.investigador_principal,
        participantes_adicionais=dados.participantes_adicionais,
        telefone_contacto=dados.telefone_contacto,
        area_tecnologica=dados.area_tecnologica, data_inicio=dados.data_inicio,
        status_trl=dados.status_trl, codigo_patente_previsto=dados.codigo_patente_previsto,
        criado_em=agora, actualizado_em=agora,
    )
    db.add(c); db.commit()
    return {"id_caderno": c.id_caderno, "codigo_ltdp": c.codigo_ltdp,
            "mensagem": f"Caderno aberto no {LOCALIZACAO}.", "registado_por": u.email}


@app.get("/cadernos", tags=["Cadernos"])
def listar_cadernos(area: Optional[str] = Query(None), trl_min: Optional[int] = Query(None, ge=1, le=9),
                    estado: Optional[str] = Query(None), db: Session = Depends(get_db),
                    _: Utilizador = Depends(utilizador_actual)):
    """Painel do Diretor: lista todos os projectos e o estado de desenvolvimento (TRL)."""
    q = db.query(Caderno)
    if area:
        q = q.filter(Caderno.area_tecnologica.ilike(f"%{area}%"))
    if trl_min:
        q = q.filter(Caderno.status_trl >= trl_min)
    if estado:
        q = q.filter(Caderno.estado == estado.upper())
    cadernos = q.order_by(Caderno.status_trl.desc()).all()
    resposta = []
    for c in cadernos:
        n = db.query(func.count(EntradaCientifica.id_entrada)).filter(
            EntradaCientifica.id_caderno == c.id_caderno).scalar()
        ultima = db.query(func.max(EntradaCientifica.data_registo)).filter(
            EntradaCientifica.id_caderno == c.id_caderno).scalar()
        resposta.append({
            "id_caderno": c.id_caderno, "codigo_ltdp": c.codigo_ltdp,
            "titulo_invencao": c.titulo_invencao,
            "investigador_principal": c.investigador_principal,
            "participantes_adicionais": c.participantes_adicionais,
            "telefone_contacto": c.telefone_contacto,
            "area_tecnologica": c.area_tecnologica, "data_inicio": c.data_inicio,
            "status_trl": c.status_trl, "trl_descricao": f"TRL {c.status_trl}/9",
            "codigo_patente_previsto": c.codigo_patente_previsto, "estado": c.estado,
            "total_entradas": n, "ultima_actividade": ultima,
        })
    return {"localizacao": LOCALIZACAO, "total": len(resposta), "cadernos": resposta}


@app.get("/cadernos/{id_caderno}/entradas", tags=["Entradas Científicas"])
def listar_entradas(id_caderno: str, db: Session = Depends(get_db),
                    _: Utilizador = Depends(utilizador_actual)):
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    entradas = (db.query(EntradaCientifica)
                .filter(EntradaCientifica.id_caderno == id_caderno)
                .order_by(EntradaCientifica.data_registo).all())
    return [{
        "id_entrada": e.id_entrada, "data_registo": e.data_registo,
        "metodologia": e.metodologia, "resultados_brutos": e.resultados_brutos,
        "link_repositorio_codigo": e.link_repositorio_codigo,
        "hash_seguranca": e.hash_seguranca, "hash_anterior": e.hash_anterior,
        "assinatura_digital_investigador": e.assinatura_digital_investigador,
        "assinatura_testemunha": e.assinatura_testemunha,
        "origem_registo": e.origem_registo,
        "anulada": bool(e.anulada or 0),
        "editada_em": e.editada_em, "editada_por": e.editada_por,
        "motivo_alteracao": e.motivo_alteracao,
    } for e in entradas]


@app.post("/cadernos/{id_caderno}/entradas", tags=["Entradas Científicas"], status_code=201)
def adicionar_entrada(id_caderno: str, dados: EntradaIn, db: Session = Depends(get_db),
                      u: Utilizador = Depends(utilizador_actual)):
    """
    Regista um dia de progresso científico. O Hub:
    1. Gera o SHA-256 canónico do conteúdo (integridade intelectual);
    2. Encadeia com o hash da entrada anterior (cadeia de auditoria);
    3. Assina digitalmente em nome do investigador autenticado.
    Idempotente: se o id_entrada (UUID offline) já existir, devolve 409 sem duplicar.
    """
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    eid = dados.id_entrada or str(uuid.uuid4())
    if db.get(EntradaCientifica, eid):
        raise HTTPException(409, "Entrada já sincronizada no Hub (sem duplicação).")

    data_registo = dados.data_registo or agora_wat().isoformat()
    h = calcular_hash_entrada(id_caderno, data_registo, dados.metodologia,
                              dados.resultados_brutos, dados.link_repositorio_codigo)
    if db.query(EntradaCientifica).filter(EntradaCientifica.hash_seguranca == h).first():
        raise HTTPException(409, "Conteúdo idêntico já registado (hash duplicado).")

    anterior = (db.query(EntradaCientifica)
                .filter(EntradaCientifica.id_caderno == id_caderno)
                .order_by(EntradaCientifica.criado_em.desc()).first())

    e = EntradaCientifica(
        id_entrada=eid, id_caderno=id_caderno, data_registo=data_registo,
        metodologia=dados.metodologia, resultados_brutos=dados.resultados_brutos,
        link_repositorio_codigo=dados.link_repositorio_codigo,
        hash_seguranca=h, hash_anterior=anterior.hash_seguranca if anterior else None,
        assinatura_digital_investigador=assinar(u.email, h),
        assinatura_testemunha=(assinar(dados.assinatura_testemunha_email, h)
                               if dados.assinatura_testemunha_email else None),
        origem_registo=dados.origem_registo, criado_em=agora_wat().isoformat(),
    )
    caderno = db.get(Caderno, id_caderno)
    caderno.actualizado_em = agora_wat().isoformat()
    db.add(e); db.commit()
    return {"id_entrada": e.id_entrada, "hash_seguranca": e.hash_seguranca,
            "hash_anterior": e.hash_anterior,
            "assinatura_digital_investigador": e.assinatura_digital_investigador,
            "localizacao": LOCALIZACAO, "fuso_horario": FUSO_LABEL}


@app.get("/entradas/{id_entrada}/verificar", tags=["Auditoria"])
def verificar_integridade(id_entrada: str, db: Session = Depends(get_db),
                          _: Utilizador = Depends(utilizador_actual)):
    """Recalcula o SHA-256 e compara com o registado — prova de não-adulteração."""
    e = db.get(EntradaCientifica, id_entrada)
    if not e:
        raise HTTPException(404, "Entrada não encontrada.")
    recalculado = calcular_hash_entrada(e.id_caderno, e.data_registo, e.metodologia,
                                        e.resultados_brutos, e.link_repositorio_codigo)
    integra = hmac.compare_digest(recalculado, e.hash_seguranca)
    return {"id_entrada": id_entrada, "hash_registado": e.hash_seguranca,
            "hash_recalculado": recalculado, "integridade": "ÍNTEGRA" if integra else "⚠ QUEBRA DE INTEGRIDADE",
            "verificado_em": agora_wat().isoformat()}


# --------------------- Edição/Anulação de Entradas --------------------------
# Reservado ao DIRETOR (administrador). Mantém valor probatório: a edição é
# auditada (quem/quando/porquê) e a eliminação é LÓGICA (soft-delete), de modo
# a preservar a cadeia de anterioridade para efeitos de patente (IAPI).
# ----------------------------------------------------------------------------
@app.put("/entradas/{id_entrada}", tags=["Entradas Científicas"])
def editar_entrada(id_entrada: str, dados: EntradaEditar, db: Session = Depends(get_db),
                   admin: Utilizador = Depends(exigir_diretor)):
    """Corrige uma entrada já gravada. Recalcula o selo SHA-256 e regista a auditoria."""
    e = db.get(EntradaCientifica, id_entrada)
    if not e:
        raise HTTPException(404, "Entrada não encontrada.")
    if dados.metodologia is not None:
        e.metodologia = dados.metodologia
    if dados.resultados_brutos is not None:
        e.resultados_brutos = dados.resultados_brutos
    if dados.link_repositorio_codigo is not None:
        e.link_repositorio_codigo = dados.link_repositorio_codigo
    if dados.data_registo is not None:
        e.data_registo = dados.data_registo

    novo_hash = calcular_hash_entrada(e.id_caderno, e.data_registo, e.metodologia,
                                      e.resultados_brutos, e.link_repositorio_codigo)
    conflito = (db.query(EntradaCientifica)
                .filter(EntradaCientifica.hash_seguranca == novo_hash,
                        EntradaCientifica.id_entrada != e.id_entrada).first())
    if conflito:
        raise HTTPException(409, "Já existe outra entrada com conteúdo idêntico (hash duplicado).")

    e.hash_seguranca = novo_hash
    e.assinatura_digital_investigador = assinar(admin.email, novo_hash)
    e.editada_em = agora_wat().isoformat()
    e.editada_por = admin.email
    e.motivo_alteracao = dados.motivo_alteracao
    db.commit()
    return {"id_entrada": e.id_entrada, "novo_hash": e.hash_seguranca,
            "editada_por": e.editada_por, "editada_em": e.editada_em,
            "aviso": "Selo SHA-256 recalculado. A correcção fica auditada."}


@app.delete("/entradas/{id_entrada}", tags=["Entradas Científicas"])
def anular_entrada(id_entrada: str, motivo: str = Query(..., min_length=3),
                   fisico: bool = Query(False, description="True = eliminação física definitiva (irreversível)."),
                   db: Session = Depends(get_db), admin: Utilizador = Depends(exigir_diretor)):
    """
    Anula (soft-delete) uma entrada gravada por erro. Por omissão a entrada é
    apenas MARCADA como anulada (preserva a cadeia de auditoria). Use fisico=true
    apenas se quiser eliminar definitivamente o registo da base de dados.
    """
    e = db.get(EntradaCientifica, id_entrada)
    if not e:
        raise HTTPException(404, "Entrada não encontrada.")
    if fisico:
        db.delete(e); db.commit()
        return {"id_entrada": id_entrada, "eliminada": "DEFINITIVA",
                "por": admin.email, "motivo": motivo}
    e.anulada = 1
    e.editada_em = agora_wat().isoformat()
    e.editada_por = admin.email
    e.motivo_alteracao = f"[ANULADA] {motivo}"
    db.commit()
    return {"id_entrada": id_entrada, "estado": "ANULADA (soft-delete)",
            "por": admin.email, "motivo": motivo}


# --------------------------- Imagens dos inventos ---------------------------
@app.post("/cadernos/{id_caderno}/imagens", tags=["Imagens"], status_code=201)
async def carregar_imagem(id_caderno: str, ficheiro: UploadFile = File(...),
                          legenda: Optional[str] = Form(None),
                          db: Session = Depends(get_db),
                          u: Utilizador = Depends(utilizador_actual)):
    """Carrega uma imagem do invento a partir do PC (ficheiro). Guardada embebida (base64)."""
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    if not (ficheiro.content_type or "").startswith("image/"):
        raise HTTPException(400, "O ficheiro tem de ser uma imagem.")
    conteudo = await ficheiro.read()
    if len(conteudo) > 5 * 1024 * 1024:
        raise HTTPException(413, "Imagem demasiado grande (máx. 5 MB).")
    import base64
    img = ImagemInvento(
        id_imagem=str(uuid.uuid4()), id_caderno=id_caderno,
        nome_ficheiro=ficheiro.filename or "imagem", tipo_mime=ficheiro.content_type,
        legenda=legenda, dados_base64=base64.b64encode(conteudo).decode("ascii"),
        carregado_por=u.email, criado_em=agora_wat().isoformat(),
    )
    db.add(img); db.commit()
    return {"id_imagem": img.id_imagem, "nome_ficheiro": img.nome_ficheiro,
            "legenda": img.legenda, "mensagem": "Imagem do invento carregada."}


@app.get("/cadernos/{id_caderno}/imagens", tags=["Imagens"])
def listar_imagens(id_caderno: str, db: Session = Depends(get_db),
                   _: Utilizador = Depends(utilizador_actual)):
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    imgs = db.query(ImagemInvento).filter(ImagemInvento.id_caderno == id_caderno).all()
    return [{"id_imagem": i.id_imagem, "nome_ficheiro": i.nome_ficheiro,
             "tipo_mime": i.tipo_mime, "legenda": i.legenda,
             "data_uri": f"data:{i.tipo_mime};base64,{i.dados_base64}",
             "carregado_por": i.carregado_por, "criado_em": i.criado_em} for i in imgs]


@app.delete("/imagens/{id_imagem}", tags=["Imagens"])
def apagar_imagem(id_imagem: str, db: Session = Depends(get_db),
                  u: Utilizador = Depends(utilizador_actual)):
    img = db.get(ImagemInvento, id_imagem)
    if not img:
        raise HTTPException(404, "Imagem não encontrada.")
    db.delete(img); db.commit()
    return {"id_imagem": id_imagem, "eliminada": True}


# ----------------------- Custos da invenção --------------------------------
@app.post("/cadernos/{id_caderno}/custos", tags=["Custos"], status_code=201)
def adicionar_custo(id_caderno: str, dados: CustoIn, db: Session = Depends(get_db),
                    u: Utilizador = Depends(utilizador_actual)):
    """Regista um custo da invenção (do início à finalização)."""
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    custo = CustoInvento(
        id_custo=str(uuid.uuid4()), id_caderno=id_caderno, descricao=dados.descricao,
        categoria=dados.categoria, valor=f"{dados.valor:.2f}", moeda=dados.moeda,
        data_despesa=dados.data_despesa, fase=dados.fase,
        registado_por=u.email, criado_em=agora_wat().isoformat(),
    )
    db.add(custo); db.commit()
    return {"id_custo": custo.id_custo, "mensagem": "Custo registado."}


@app.get("/cadernos/{id_caderno}/custos", tags=["Custos"])
def listar_custos(id_caderno: str, db: Session = Depends(get_db),
                  _: Utilizador = Depends(utilizador_actual)):
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    custos = db.query(CustoInvento).filter(CustoInvento.id_caderno == id_caderno).order_by(
        CustoInvento.data_despesa).all()
    lista = [{"id_custo": c.id_custo, "descricao": c.descricao, "categoria": c.categoria,
              "valor": float(c.valor), "moeda": c.moeda, "data_despesa": c.data_despesa,
              "fase": c.fase, "registado_por": c.registado_por} for c in custos]
    total = sum(x["valor"] for x in lista)
    por_fase, por_categoria = {}, {}
    for x in lista:
        por_fase[x["fase"]] = por_fase.get(x["fase"], 0) + x["valor"]
        por_categoria[x["categoria"]] = por_categoria.get(x["categoria"], 0) + x["valor"]
    moeda = lista[0]["moeda"] if lista else "AOA"
    return {"custos": lista, "total": round(total, 2), "moeda": moeda,
            "total_por_fase": {k: round(v, 2) for k, v in por_fase.items()},
            "total_por_categoria": {k: round(v, 2) for k, v in por_categoria.items()}}


@app.delete("/custos/{id_custo}", tags=["Custos"])
def apagar_custo(id_custo: str, db: Session = Depends(get_db),
                 u: Utilizador = Depends(utilizador_actual)):
    c = db.get(CustoInvento, id_custo)
    if not c:
        raise HTTPException(404, "Custo não encontrado.")
    db.delete(c); db.commit()
    return {"id_custo": id_custo, "eliminado": True}


# ---------------- Importação de dados a partir de Excel (PC) ----------------
@app.post("/cadernos/{id_caderno}/importar/excel", tags=["Importação"])
async def importar_entradas_excel(id_caderno: str, ficheiro: UploadFile = File(...),
                                  db: Session = Depends(get_db),
                                  u: Utilizador = Depends(utilizador_actual)):
    """
    Importa entradas científicas a partir de um ficheiro Excel (.xlsx) do PC.
    Colunas esperadas (linha 1 = cabeçalho), nesta ordem ou por nome:
      Data Registo | Metodologia | Resultados Brutos | Link Repositorio (opcional)
    Idempotente: linhas cujo conteúdo gera um hash já existente são ignoradas.
    """
    if not db.get(Caderno, id_caderno):
        raise HTTPException(404, "Caderno não encontrado.")
    if not (ficheiro.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Envie um ficheiro Excel (.xlsx).")
    from openpyxl import load_workbook
    conteudo = await ficheiro.read()
    try:
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception:
        raise HTTPException(400, "Não foi possível ler o ficheiro Excel.")
    ws = wb.active

    # Mapeia cabeçalhos (tolerante a maiúsculas/acentos básicos)
    def norm(s):
        return (str(s or "")).strip().lower()
    cabec = [norm(c.value) for c in ws[1]] if ws.max_row >= 1 else []
    def col(*nomes):
        for n in nomes:
            if n in cabec:
                return cabec.index(n)
        return None
    i_data = col("data registo", "data_registo", "data")
    i_met = col("metodologia")
    i_res = col("resultados brutos", "resultados_brutos", "resultados")
    i_link = col("link repositorio", "link_repositorio", "link repositório", "link")

    novas, ignoradas, erros = 0, 0, []
    for n_linha, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def get(idx):
            return linha[idx] if idx is not None and idx < len(linha) else None
        metodologia = get(i_met) if i_met is not None else (linha[1] if len(linha) > 1 else None)
        resultados = get(i_res) if i_res is not None else (linha[2] if len(linha) > 2 else None)
        if not metodologia and not resultados:
            continue
        data_registo = get(i_data) if i_data is not None else (linha[0] if linha else None)
        if hasattr(data_registo, "isoformat"):
            data_registo = data_registo.isoformat()
        data_registo = str(data_registo) if data_registo else agora_wat().isoformat()
        link = get(i_link)
        link = str(link) if link else None
        metodologia = str(metodologia or "").strip()
        resultados = str(resultados or "").strip()
        if not metodologia or not resultados:
            erros.append(f"Linha {n_linha}: metodologia/resultados em falta.")
            continue
        h = calcular_hash_entrada(id_caderno, data_registo, metodologia, resultados, link)
        if db.query(EntradaCientifica).filter(EntradaCientifica.hash_seguranca == h).first():
            ignoradas += 1
            continue
        anterior = (db.query(EntradaCientifica)
                    .filter(EntradaCientifica.id_caderno == id_caderno)
                    .order_by(EntradaCientifica.criado_em.desc()).first())
        db.add(EntradaCientifica(
            id_entrada=str(uuid.uuid4()), id_caderno=id_caderno, data_registo=data_registo,
            metodologia=metodologia, resultados_brutos=resultados,
            link_repositorio_codigo=link, hash_seguranca=h,
            hash_anterior=anterior.hash_seguranca if anterior else None,
            assinatura_digital_investigador=assinar(u.email, h),
            origem_registo="EXCEL_SYNC", criado_em=agora_wat().isoformat(),
        ))
        novas += 1
    caderno = db.get(Caderno, id_caderno)
    caderno.actualizado_em = agora_wat().isoformat()
    db.add(LogSincronizacao(
        id_sync=str(uuid.uuid4()), id_utilizador=u.id_utilizador,
        data_sync=agora_wat().isoformat(), entradas_novas=novas, duplicados=ignoradas,
        alertas_hash=0, ficheiro_origem=ficheiro.filename,
        detalhes=json.dumps({"erros": erros}, ensure_ascii=False),
    ))
    db.commit()
    return {"ficheiro": ficheiro.filename, "entradas_novas": novas,
            "ignoradas_duplicadas": ignoradas, "erros": erros}


# --------------------------- Exportação Excel -------------------------------
@app.get("/cadernos/export/excel", tags=["Exportação"])
def exportar_excel(db: Session = Depends(get_db), _: Utilizador = Depends(exigir_diretor)):
    """Dump completo formatado (.xlsx) para arquivo/relatório do Diretor."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    azul = PatternFill("solid", start_color="1F4E79")
    cab = Font(bold=True, color="FFFFFF", name="Arial")

    ws1 = wb.active
    ws1.title = "Cadernos"
    cols1 = ["ID_Caderno", "Código LTDP", "Título da Invenção", "Investigador Principal",
             "Outros Participantes", "Telefone", "Área Tecnológica", "Data Início",
             "Status TRL", "Código Patente Previsto", "Estado", "Localização", "Fuso Horário"]
    ws1.append(cols1)
    for c in db.query(Caderno).all():
        ws1.append([c.id_caderno, c.codigo_ltdp, c.titulo_invencao, c.investigador_principal,
                    c.participantes_adicionais, c.telefone_contacto,
                    c.area_tecnologica, c.data_inicio, c.status_trl, c.codigo_patente_previsto,
                    c.estado, c.localizacao, c.fuso_horario])

    ws2 = wb.create_sheet("Entradas_Cientificas")
    cols2 = ["ID_Entrada", "ID_Caderno", "Data Registo", "Metodologia", "Resultados Brutos",
             "Link Repositório", "Hash SHA-256", "Hash Anterior", "Assinatura Investigador",
             "Assinatura Testemunha", "Origem", "Anulada", "Editada Por", "Motivo Alteração"]
    ws2.append(cols2)
    for e in db.query(EntradaCientifica).order_by(EntradaCientifica.data_registo).all():
        ws2.append([e.id_entrada, e.id_caderno, e.data_registo, e.metodologia,
                    e.resultados_brutos, e.link_repositorio_codigo, e.hash_seguranca,
                    e.hash_anterior, e.assinatura_digital_investigador,
                    e.assinatura_testemunha, e.origem_registo,
                    "SIM" if (e.anulada or 0) else "", e.editada_por or "", e.motivo_alteracao or ""])

    ws3 = wb.create_sheet("Custos")
    cols3 = ["ID_Custo", "ID_Caderno", "Descrição", "Categoria", "Valor", "Moeda",
             "Data Despesa", "Fase", "Registado Por"]
    ws3.append(cols3)
    for k in db.query(CustoInvento).order_by(CustoInvento.id_caderno, CustoInvento.data_despesa).all():
        ws3.append([k.id_custo, k.id_caderno, k.descricao, k.categoria, float(k.valor),
                    k.moeda, k.data_despesa, k.fase, k.registado_por])

    for ws, cols in ((ws1, cols1), (ws2, cols2), (ws3, cols3)):
        for i, _c in enumerate(cols, 1):
            cel = ws.cell(row=1, column=i)
            cel.fill, cel.font = azul, cab
            cel.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cel.column_letter].width = 24
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"LTDP_Hub_Export_{agora_wat().strftime('%Y%m%d_%H%M')}_WAT.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )
