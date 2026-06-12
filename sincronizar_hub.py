# ============================================================================
# MOTOR DE SINCRONIZAÇÃO EXCEL <-> HUB LTDP
# Laboratório Tecnológico de Desenvolvimento e Pesquisa
# Lubango, Província da Huíla, Angola — WAT (UTC+1)
# ----------------------------------------------------------------------------
# Fluxo de trabalho do investigador:
#   1. Trabalha OFFLINE na planilha "Caderno_Cientifico_LTDP.xlsx" durante o dia;
#   2. Quando há internet no laboratório, executa:
#        python sincronizar_hub.py Caderno_Cientifico_LTDP.xlsx
#      (ou clica no botão "Sincronizar com o Hub LTDP" na planilha, que invoca
#       este script via macro/atalho do sistema);
#   3. O script envia apenas as linhas novas, marca cada uma como
#      SINCRONIZADO, devolve o Hash SHA-256 oficial do Hub e gera um
#      relatório de conflitos/integridade.
#
# Regras de conflito:
#   - Linha com ID_Entrada já existente na API  -> NÃO duplica (idempotente);
#   - Hash local preenchido que difere do recálculo -> ALERTA DE QUEBRA DE
#     INTEGRIDADE (a linha foi alterada após o registo) e a linha NÃO é enviada;
#   - Falha de rede a meio -> as linhas já confirmadas ficam marcadas; as
#     restantes permanecem PENDENTE e serão enviadas na próxima sincronização.
# ============================================================================

import hashlib
import json
import sys
import uuid
from datetime import datetime, timedelta, timezone
from getpass import getpass

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ----------------------------- Configuração ---------------------------------
import os
API_URL = os.getenv("LTDP_API_URL", "http://localhost:8000").rstrip("/")  # ex.: https://ltdp-hub.onrender.com
FOLHA_CADERNO = "Caderno"
FOLHA_ENTRADAS = "Entradas_Cientificas"
LINHA_INICIO = 5                         # primeira linha de dados na folha de entradas
LOCALIZACAO = "LTDP - Lubango, Huíla, Angola"
WAT = timezone(timedelta(hours=1), name="WAT")

# Mapa de colunas da folha "Entradas_Cientificas" (A=1 ... I=9)
COL = {
    "id_entrada": 1, "data_registo": 2, "metodologia": 3, "resultados_brutos": 4,
    "link_repo": 5, "hash": 6, "ass_investigador": 7, "ass_testemunha": 8, "estado": 9,
}

VERDE = PatternFill("solid", start_color="C6EFCE")
VERMELHO = PatternFill("solid", start_color="FFC7CE")
AMARELO = PatternFill("solid", start_color="FFEB9C")


def calcular_hash_entrada(id_caderno, data_registo, metodologia, resultados, link):
    """Réplica EXACTA da fórmula canónica do backend (main.py)."""
    payload = json.dumps(
        {
            "id_caderno": id_caderno,
            "data_registo": data_registo,
            "metodologia": str(metodologia).strip(),
            "resultados_brutos": str(resultados).strip(),
            "link_repositorio_codigo": str(link or "").strip(),
            "contexto": LOCALIZACAO,
        },
        ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def autenticar(sessao: requests.Session) -> None:
    print(f"\n=== Sincronização com o Hub LTDP ({LOCALIZACAO}) ===")
    email = input("Email LTDP: ").strip()
    senha = getpass("Senha: ")
    r = sessao.post(f"{API_URL}/auth/login", data={"username": email, "password": senha}, timeout=30)
    if r.status_code != 200:
        sys.exit("✗ Autenticação falhou. Verifique as credenciais.")
    sessao.headers["Authorization"] = f"Bearer {r.json()['access_token']}"
    print("✓ Sessão JWT iniciada (válida 12h — adequada a redes instáveis).")


def garantir_caderno(sessao, ws_caderno) -> str:
    """Lê os metadados da folha 'Caderno'; cria o caderno no Hub se ainda não existir."""
    meta = {
        "id_caderno": ws_caderno["B2"].value,
        "titulo_invencao": ws_caderno["B3"].value,
        "investigador_principal": ws_caderno["B4"].value,
        "area_tecnologica": ws_caderno["B5"].value,
        "data_inicio": str(ws_caderno["B6"].value)[:10],
        "status_trl": int(ws_caderno["B7"].value or 1),
        "codigo_patente_previsto": ws_caderno["B8"].value,
    }
    if not meta["titulo_invencao"]:
        sys.exit("✗ A folha 'Caderno' não tem o título da invenção preenchido (célula B3).")

    if not meta["id_caderno"]:
        meta["id_caderno"] = str(uuid.uuid4())
        ws_caderno["B2"] = meta["id_caderno"]

    r = sessao.post(f"{API_URL}/cadernos", json=meta, timeout=30)
    if r.status_code == 201:
        ws_caderno["B9"] = r.json()["codigo_ltdp"]
        print(f"✓ Caderno aberto no Hub: {r.json()['codigo_ltdp']}")
    elif r.status_code == 409:
        print("• Caderno já existente no Hub — a sincronizar apenas as entradas.")
    else:
        sys.exit(f"✗ Erro ao registar o caderno: {r.status_code} {r.text}")
    return meta["id_caderno"]


def sincronizar(caminho_xlsx: str) -> None:
    wb = load_workbook(caminho_xlsx)
    ws_cad, ws_ent = wb[FOLHA_CADERNO], wb[FOLHA_ENTRADAS]

    sessao = requests.Session()
    try:
        sessao.get(f"{API_URL}/health", timeout=30).raise_for_status()
    except requests.RequestException:
        sys.exit("✗ Sem ligação ao Hub LTDP. Continue a trabalhar offline e "
                 "sincronize quando a internet do laboratório estiver disponível.")

    autenticar(sessao)
    id_caderno = garantir_caderno(sessao, ws_cad)

    novas, duplicadas, alertas, erros = 0, 0, 0, 0

    for linha in range(LINHA_INICIO, ws_ent.max_row + 1):
        metodologia = ws_ent.cell(linha, COL["metodologia"]).value
        resultados = ws_ent.cell(linha, COL["resultados_brutos"]).value
        if not metodologia and not resultados:
            continue  # linha vazia

        estado = (ws_ent.cell(linha, COL["estado"]).value or "").strip().upper()
        if estado == "SINCRONIZADO":
            # ----- Verificação de integridade pós-sincronização -----
            hash_local = recalcular_linha(ws_ent, linha, id_caderno)
            hash_guardado = ws_ent.cell(linha, COL["hash"]).value
            if hash_guardado and hash_local != hash_guardado:
                marcar(ws_ent, linha, "⚠ QUEBRA DE INTEGRIDADE", VERMELHO)
                print(f"⚠ ALERTA linha {linha}: o conteúdo foi alterado após o registo "
                      f"no Hub — o Hash SHA-256 já não corresponde. Para corrigir um "
                      f"registo, crie uma NOVA entrada (princípio de caderno imutável).")
                alertas += 1
            continue

        # ----- Preparar nova entrada -----
        id_entrada = ws_ent.cell(linha, COL["id_entrada"]).value or str(uuid.uuid4())
        ws_ent.cell(linha, COL["id_entrada"], id_entrada)

        data_registo = ws_ent.cell(linha, COL["data_registo"]).value
        if isinstance(data_registo, datetime):
            data_registo = data_registo.replace(tzinfo=WAT).isoformat()
        elif not data_registo:
            data_registo = datetime.now(WAT).isoformat()
            ws_ent.cell(linha, COL["data_registo"], data_registo)
        else:
            data_registo = str(data_registo)

        corpo = {
            "id_entrada": id_entrada,
            "data_registo": data_registo,
            "metodologia": str(metodologia or ""),
            "resultados_brutos": str(resultados or ""),
            "link_repositorio_codigo": ws_ent.cell(linha, COL["link_repo"]).value,
            "origem_registo": "EXCEL_SYNC",
        }

        try:
            r = sessao.post(f"{API_URL}/cadernos/{id_caderno}/entradas", json=corpo, timeout=30)
        except requests.RequestException:
            marcar(ws_ent, linha, "PENDENTE (rede caiu)", AMARELO)
            erros += 1
            print(f"✗ Linha {linha}: a ligação caiu — ficará pendente para a próxima sincronização.")
            continue

        if r.status_code == 201:
            dados = r.json()
            ws_ent.cell(linha, COL["hash"], dados["hash_seguranca"])
            ws_ent.cell(linha, COL["ass_investigador"], dados["assinatura_digital_investigador"])
            marcar(ws_ent, linha, "SINCRONIZADO", VERDE)
            novas += 1
        elif r.status_code == 409:
            marcar(ws_ent, linha, "SINCRONIZADO", VERDE)  # já existe no Hub — não duplica
            duplicadas += 1
        else:
            marcar(ws_ent, linha, f"ERRO {r.status_code}", VERMELHO)
            erros += 1

    wb.save(caminho_xlsx)
    agora = datetime.now(WAT).strftime("%d/%m/%Y %H:%M WAT")
    print(f"\n=== Relatório de Sincronização — {agora} ===")
    print(f"  Novas entradas enviadas : {novas}")
    print(f"  Duplicados ignorados    : {duplicadas}")
    print(f"  Alertas de integridade  : {alertas}")
    print(f"  Erros/pendentes         : {erros}")
    print(f"  Planilha actualizada    : {caminho_xlsx}")


def recalcular_linha(ws, linha, id_caderno):
    data = ws.cell(linha, COL["data_registo"]).value
    if isinstance(data, datetime):
        data = data.replace(tzinfo=WAT).isoformat()
    return calcular_hash_entrada(
        id_caderno, str(data),
        ws.cell(linha, COL["metodologia"]).value or "",
        ws.cell(linha, COL["resultados_brutos"]).value or "",
        ws.cell(linha, COL["link_repo"]).value,
    )


def marcar(ws, linha, texto, cor):
    cel = ws.cell(linha, COL["estado"], texto)
    cel.fill = cor
    cel.font = Font(bold=True, name="Arial", size=9)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python sincronizar_hub.py <Caderno_Cientifico_LTDP.xlsx>")
    sincronizar(sys.argv[1])
